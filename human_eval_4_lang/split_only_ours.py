import os
import json
import random
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Side, PatternFill
from PIL import Image as PILImage

# ================= 🔴 配置区域 🔴 =================

# 只需要读取最终生成的“Ours”数据文件夹
FINAL_DATA_DIR = r"E:\桌面\XJUNLP\Construction Dataset\dataset_code\all_pipline\10_final_datasets_split"
# 图片文件夹保持不变
IMAGE_ROOT_DIR = r"E:\桌面\XJUNLP\Construction Dataset\ImageNet50K\ILSVRC2012_img_val"
OUTPUT_DIR = r"E:\桌面\XJUNLP\Construction Dataset\dataset_code\human_eval_single_model"

# 修改1：只保留指定的4种语言
LANGUAGES = ['kazakh', 'urdu', 'uyghur', 'uzbek']

# 修改2：只保留 short 类型
TYPES = ['short']

# 采样数量
SAMPLE_SIZE =  250

# 随机种子
RANDOM_SEED = 42


# =======================================================

def load_json(path):
    if not os.path.exists(path):
        print(f"❌ [Error] File not found: {path}")
        return None
    print(f"📖 Loading: {os.path.basename(path)} ...")
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"❌ [Error] JSON Load Failed: {e}")
        return None


def resize_image(image_path, save_path, max_size=(400, 400)):
    """
    调整图片大小，适应单行高度
    """
    try:
        if not os.path.exists(image_path): return False
        with PILImage.open(image_path) as img:
            if img.mode != 'RGB': img = img.convert('RGB')
            img.thumbnail(max_size, resample=PILImage.LANCZOS)
            img.save(save_path, quality=95)
        return True
    except Exception as e:
        return False


def main():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    print(f"🎲 Setting Random Seed to: {RANDOM_SEED}")
    random.seed(RANDOM_SEED)

    for lang in LANGUAGES:
        print(f"Processing Language: {lang.upper()} ...")
        wb = Workbook()
        if 'Sheet' in wb.sheetnames: del wb['Sheet']

        for cap_type in TYPES:
            final_file = os.path.join(FINAL_DATA_DIR, f"{lang}_{cap_type}.json")
            final_data = load_json(final_file)

            if not final_data:
                print(f"⚠️ Skipping {lang}_{cap_type}: Data file not found.")
                continue

            # 随机抽样
            current_sample_size = min(SAMPLE_SIZE, len(final_data))
            samples = random.sample(final_data, current_sample_size)

            ws = wb.create_sheet(title=cap_type.capitalize())

            # 修改3：表头简化，去掉 Model 对比列，去掉 Secret Key
            headers = [
                "Image",  # A列
                "ID",  # B列
                "English Source",  # C列
                "Translation",  # D列 (Ours)
                "Fluency",  # E列
                "Adequacy",  # F列
                "Visual Relevance"  # G列
            ]
            ws.append(headers)

            for idx, item in enumerate(samples):
                img_id = item.get('image_id')
                json_path = item.get('path', '')
                local_img_path = os.path.join(IMAGE_ROOT_DIR, os.path.basename(json_path))

                src_text = item.get('src_text', '')
                text_ours = item.get('tgt_text', '')

                # 现在的行号 (标题占1行，数据从第2行开始，每条数据只占1行)
                current_row = idx + 2

                # 写入数据
                ws.cell(row=current_row, column=2, value=img_id)
                ws.cell(row=current_row, column=3, value=src_text)
                ws.cell(row=current_row, column=4, value=text_ours)

                # 插入图片
                temp_img_name = f"tmp_{lang}_{cap_type}_{idx}.jpg"
                # 调整图片大小以适应单行
                if resize_image(local_img_path, temp_img_name, max_size=(380, 380)):
                    img = XLImage(temp_img_name)
                    # 稍微留点边距
                    ws.add_image(img, f"A{current_row}")
                else:
                    ws.cell(row=current_row, column=1, value="Img Not Found")

                # 设置行高 (设置为300大约对应400像素的高度，足以容纳图片)
                ws.row_dimensions[current_row].height = 300

            # 样式美化
            ws.column_dimensions['A'].width = 55  # 图片列宽
            ws.column_dimensions['B'].width = 15  # ID
            ws.column_dimensions['C'].width = 40  # 英文源
            ws.column_dimensions['D'].width = 50  # 译文
            ws.column_dimensions['E'].width = 10  # 打分列
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 15

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            align = Alignment(wrap_text=True, vertical='center', horizontal='left')

            # 应用样式
            max_row = 1 + len(samples)
            for row in ws.iter_rows(min_row=1, max_row=max_row):
                for cell in row:
                    cell.alignment = align
                    cell.border = thin_border
                    # 表头加粗居中
                    if cell.row == 1:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        save_path = os.path.join(OUTPUT_DIR, f"Human_Eval_Short_{lang.capitalize()}.xlsx")
        while True:
            try:
                wb.save(save_path)
                print(f"✅ Saved: {save_path}")
                break
            except PermissionError:
                input("❌ Excel文件被占用！请关闭Excel后按回车重试...")
            except Exception as e:
                print(e)
                break

    print("🧹 Cleaning temp files...")
    for f in os.listdir("."):
        if f.startswith("tmp_") and f.endswith(".jpg"):
            try:
                os.remove(f)
            except:
                pass
    print("🎉 Done! Random Seed was:", RANDOM_SEED)


if __name__ == "__main__":
    main()